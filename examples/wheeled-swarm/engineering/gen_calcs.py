#!/usr/bin/env python3
"""
Generate swarm-robot-calcs.xlsx — engineering worksheet for the differential-drive
swarm robot: weight, drive (speed + torque margin), and power/runtime, with LIVE
formulas. Yellow cells are inputs (edit for your parts); everything else updates.
Starting values are realistic for a TT-gear-motor build — VERIFY against your parts.

    pip install openpyxl
    python gen_calcs.py        # -> swarm-robot-calcs.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Weight-Drive-Power"
IN = PatternFill("solid", fgColor="FFF2C2"); HDR = PatternFill("solid", fgColor="1F3B63")
TITLE = PatternFill("solid", fgColor="0B0E14"); VERD = PatternFill("solid", fgColor="EAF3EA")
thin = Side(style="thin", color="C9D3E0"); border = Border(thin, thin, thin, thin)
B = Font(bold=True); WB = Font(bold=True, color="FFFFFF")


def c(coord, v=None, *, fill=None, font=None, num=None, align=None, bd=False, wrap=False):
    cell = ws[coord]
    if v is not None: cell.value = v
    if fill: cell.fill = fill
    if font: cell.font = font
    if num: cell.number_format = num
    if align or wrap: cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")
    if bd: cell.border = border
    return cell


def section(r, t):
    ws.merge_cells(f"A{r}:D{r}"); c(f"A{r}", t, fill=HDR, font=WB, align="left")


for col, w in {"A": 40, "B": 13, "C": 8, "D": 46}.items():
    ws.column_dimensions[col].width = w
ws.merge_cells("A1:D1")
c("A1", "Differential-Drive Swarm Robot — Weight · Drive · Power",
  fill=TITLE, font=Font(bold=True, size=14, color="FFFFFF"), align="left")
ws.merge_cells("A2:D2")
c("A2", "Yellow = edit these for your parts. Everything else auto-calculates. "
  "Realistic TT-motor estimates — verify against your own parts & a quick motor test.",
  font=Font(italic=True, size=9, color="5B6B7D"), align="left")

# 1) WEIGHT
section(4, "1)  WEIGHT")
for col, t in zip("ABCD", ["Component", "Qty", "Unit g", "Total g"]):
    c(f"{col}5", t, font=B, bd=True, align="center" if col != "A" else "left")
comps = [("Chassis (3D-printed, ~40% infill)", 1, 14.0), ("TT gear motor + wheel", 2, 32.0),
         ("Motor driver (DRV8833)", 1, 2.0), ("ESP32 board", 1, 8.0),
         ("VL53L0X distance sensor", 1, 1.5), ("Ball caster", 1, 6.0),
         ("Wiring / connectors", 1, 6.0), ("ArUco tag (paper)", 1, 0.5)]
r0 = 6
for i, (name, q, um) in enumerate(comps):
    r = r0 + i
    c(f"A{r}", name, bd=True); c(f"B{r}", q, fill=IN, bd=True, align="center", num="0")
    c(f"C{r}", um, fill=IN, bd=True, align="center", num="0.0")
    c(f"D{r}", f"=B{r}*C{r}", bd=True, align="center", num="0.0")
r_dry = r0 + len(comps)
c(f"A{r_dry}", "Dry mass (no battery)", font=B, bd=True)
c(f"D{r_dry}", f"=SUM(D{r0}:D{r_dry-1})", font=B, bd=True, align="center", num="0.0")
r_b = r_dry + 1
c(f"A{r_b}", "Battery (4×AA holder + cells)", bd=True); c(f"B{r_b}", 1, fill=IN, bd=True, align="center", num="0")
c(f"C{r_b}", 100.0, fill=IN, bd=True, align="center", num="0.0"); c(f"D{r_b}", f"=B{r_b}*C{r_b}", bd=True, align="center", num="0.0")
r_auw = r_b + 1
c(f"A{r_auw}", "ALL-UP MASS", font=Font(bold=True, color="1F3B63"), bd=True)
c(f"D{r_auw}", f"=D{r_dry}+D{r_b}", font=Font(bold=True, color="1F3B63"), bd=True, align="center", num="0.0")
AUW = f"D{r_auw}"


def kv(r, lab, val, unit="", note="", inp=False, num="0.00"):
    c(f"A{r}", lab, bd=True); c(f"B{r}", val, fill=IN if inp else None, bd=True, align="center", num=num)
    c(f"C{r}", unit, bd=True, align="center")
    if note: c(f"D{r}", note, bd=True, font=Font(size=9, color="5B6B7D"), wrap=True)


# 2) DRIVE
section(17, "2)  DRIVE  (speed + torque)")
kv(18, "Motor no-load speed", 200, "rpm", "input — TT motor @ your pack voltage", inp=True, num="0")
kv(19, "Motor stall torque", 0.078, "N·m", "input — ~0.8 kg·cm for a TT motor", inp=True, num="0.000")
kv(20, "Wheel radius", 0.033, "m", "input — 66 mm wheel", inp=True, num="0.000")
kv(21, "Loaded-speed fraction", 0.5, "", "of no-load, under load", inp=True, num="0.00")
kv(22, "Rolling-resistance coeff", 0.06, "", "input — hard floor ~0.02-0.08", inp=True, num="0.00")
kv(23, "Max (no-load) speed", "=B18/60*2*PI()*B20", "m/s")
ws["B23"].number_format = "0.00"
kv(24, "Expected cruise speed", "=B23*B21", "m/s")
ws["B24"].number_format = "0.00"
kv(25, "Rolling resistance force", f"=B22*{AUW}/1000*9.81", "N")
kv(26, "Torque needed per wheel", "=B25/2*B20", "N·m", "to keep it rolling")
ws["B26"].number_format = "0.0000"
kv(27, "Torque margin (stall / needed)", "=B19/B26", "×", "how much spare torque you have")
ws["B27"].number_format = "0.0"

# 3) POWER / RUNTIME
section(29, "3)  POWER / RUNTIME")
kv(30, "Motor running current (each)", 0.20, "A", "input — TT motor, light load", inp=True)
kv(31, "Number of motors", 2, "", "", inp=True, num="0")
kv(32, "Electronics current (ESP32+ToF)", 0.12, "A", "", inp=True)
kv(33, "Battery capacity", 2000, "mAh", "input — 4×AA NiMH ~2000, LiPo varies", inp=True, num="0")
kv(34, "Usable capacity fraction", 0.8, "", "", inp=True, num="0.00")
kv(35, "Total current (driving)", "=B30*B31+B32", "A")
kv(36, "Estimated run time", "=B33*B34/(B35*1000)*60", "min")
ws["B36"].number_format = "0"

# 4) VERDICT
section(38, "4)  VERDICT")
V = [("Can it drive its own weight?",
      '=IF(B27>=3,"YES — "&TEXT(B27,"0")&"× torque margin","MARGINAL — "&TEXT(B27,"0.0")&"×; lighter build or geared motor")'),
     ("Cruise speed",
      '=TEXT(B24*100,"0")&" cm/s (plenty for a camera-tracked swarm)"'),
     ("Run time", '=TEXT(B36,"0")&" min per charge"')]
for i, (lab, f) in enumerate(V):
    r = 39 + i
    c(f"A{r}", lab, font=B, fill=VERD, bd=True); ws.merge_cells(f"B{r}:D{r}")
    c(f"B{r}", f, fill=VERD, bd=True, align="left")

ws.freeze_panes = "A3"
ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True); ws.print_area = "A1:D41"
wb.save("swarm-robot-calcs.xlsx")
print("swarm-robot-calcs.xlsx written")
