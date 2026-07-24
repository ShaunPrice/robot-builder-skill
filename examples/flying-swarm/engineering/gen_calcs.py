#!/usr/bin/env python3
"""
Generate swarm-drone-calcs.xlsx — an engineering worksheet for the ESP32 brushed
micro-quad: weight budget, motor/thrust, and power/battery, with LIVE formulas.

Yellow cells are inputs (edit them for your actual parts); everything else
recalculates. Numbers are realistic starting estimates for an 8.5 mm coreless
whoop-class build — VERIFY against your own motor test and part weights.

    pip install openpyxl
    python gen_calcs.py        # -> swarm-drone-calcs.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Motor-Weight-Power"

IN = PatternFill("solid", fgColor="FFF2C2")      # input (yellow)
HDR = PatternFill("solid", fgColor="1F3B63")      # section header
TITLE = PatternFill("solid", fgColor="0B0E14")
VERD = PatternFill("solid", fgColor="EAF3EA")
thin = Side(style="thin", color="C9D3E0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
BOLD = Font(bold=True)
WHITEB = Font(bold=True, color="FFFFFF")

def cell(coord, val=None, *, fill=None, font=None, num=None, align=None, bd=False, wrap=False):
    c = ws[coord]
    if val is not None: c.value = val
    if fill: c.fill = fill
    if font: c.font = font
    if num: c.number_format = num
    if align: c.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")
    elif wrap: c.alignment = Alignment(wrap_text=True, vertical="center")
    if bd: c.border = border
    return c

def section(row, text):
    ws.merge_cells(f"A{row}:D{row}")
    cell(f"A{row}", text, fill=HDR, font=WHITEB, align="left")

# widths
for col, w in {"A": 40, "B": 13, "C": 8, "D": 46}.items():
    ws.column_dimensions[col].width = w

# ---- title ----
ws.merge_cells("A1:D1")
cell("A1", "ESP32 Brushed Micro-Quad — Motor · Weight · Power",
     fill=TITLE, font=Font(bold=True, size=14, color="FFFFFF"), align="left")
ws.merge_cells("A2:D2")
cell("A2", "Yellow = edit these for your parts. Everything else auto-calculates. "
     "Values are realistic estimates — verify against your own motor test & part weights.",
     font=Font(italic=True, size=9, color="5B6B7D"), align="left")

# ---- 1) WEIGHT BUDGET ----
section(4, "1)  WEIGHT BUDGET")
for c, t in zip("ABCD", ["Component", "Qty", "Unit g", "Total g"]):
    cell(f"{c}5", t, font=BOLD, bd=True, align="center" if c != "A" else "left")
comps = [
    ("Frame (3D-printed, ~40% infill)", 1, 2.8),
    ("Coreless motor 8.5×20 mm", 4, 4.8),
    ("Propeller 65 mm", 4, 0.5),
    ("ESP32 board (XIAO-C3 class)", 1, 3.0),
    ("IMU MPU-6050 (trimmed module)", 1, 1.5),
    ("MOSFET driver (4× SI2302)", 1, 2.0),
    ("Wiring / connectors / heatshrink", 1, 3.0),
]
r0 = 6
for i, (name, qty, um) in enumerate(comps):
    r = r0 + i
    cell(f"A{r}", name, bd=True)
    cell(f"B{r}", qty, fill=IN, bd=True, align="center", num="0")
    cell(f"C{r}", um, fill=IN, bd=True, align="center", num="0.0")
    cell(f"D{r}", f"=B{r}*C{r}", bd=True, align="center", num="0.0")
r_dry = r0 + len(comps)          # 13
cell(f"A{r_dry}", "Dry mass (no battery)", font=BOLD, bd=True)
cell(f"D{r_dry}", f"=SUM(D{r0}:D{r_dry-1})", font=BOLD, bd=True, align="center", num="0.0")
r_batt = r_dry + 1               # 14
cell(f"A{r_batt}", "Battery — 1S LiPo 300 mAh", bd=True)
cell(f"B{r_batt}", 1, fill=IN, bd=True, align="center", num="0")
cell(f"C{r_batt}", 8.0, fill=IN, bd=True, align="center", num="0.0")
cell(f"D{r_batt}", f"=B{r_batt}*C{r_batt}", bd=True, align="center", num="0.0")
r_auw = r_batt + 1               # 15
cell(f"A{r_auw}", "ALL-UP WEIGHT (AUW)", font=Font(bold=True, color="1F3B63"), bd=True)
cell(f"D{r_auw}", f"=D{r_dry}+D{r_batt}", font=Font(bold=True, color="1F3B63"), bd=True,
     align="center", num="0.0")
AUW = f"D{r_auw}"

# ---- 2) MOTOR / THRUST ----
section(17, "2)  MOTOR / THRUST")
def kv(r, label, val, unit="", note="", inp=False, num="0.00"):
    cell(f"A{r}", label, bd=True)
    cell(f"B{r}", val, fill=IN if inp else None, bd=True, align="center", num=num)
    cell(f"C{r}", unit, bd=True, align="center")
    if note: cell(f"D{r}", note, bd=True, font=Font(size=9, color="5B6B7D"), wrap=True)
kv(18, "Motor thrust @ 100% (per motor)", 25, "g", "input — bench-test with a scale", inp=True, num="0.0")
kv(19, "Motor current @ 100% (per motor)", 1.8, "A", "input — measure at full throttle", inp=True)
kv(20, "Number of motors", 4, "", "", inp=True, num="0")
kv(21, "Total max thrust", "=B18*B20", "g")
ws["B21"].number_format = "0.0"
kv(22, "Thrust-to-weight ratio (T/W)", f"=B21/{AUW}", "", "target ≥ 2.0 : 1 for a controllable quad")
kv(23, "Hover thrust required per motor", f"={AUW}/B20", "g")
ws["B23"].number_format = "0.0"
kv(24, "Hover thrust fraction", "=B23/B18", "", "share of full thrust needed to hover")

# ---- 3) POWER / BATTERY ----
section(26, "3)  POWER / BATTERY")
kv(27, "Cells in series (S)", 1, "", "", inp=True, num="0")
kv(28, "Nominal cell voltage", 3.7, "V", "", inp=True, num="0.0")
kv(29, "Battery capacity", 300, "mAh", "input", inp=True, num="0")
kv(30, "Discharge C-rating", 30, "C", "input — high-C whoop pack, NOT a protected pack", inp=True, num="0")
kv(31, "Usable capacity fraction", 0.8, "", "never run a LiPo flat", inp=True, num="0.00")
kv(32, "Electronics current (ESP32 + IMU)", 0.15, "A", "", inp=True)
kv(33, "Pack nominal voltage", "=B27*B28", "V")
ws["B33"].number_format = "0.0"
kv(34, "Peak current the pack can supply", "=B29/1000*B30", "A", "capacity(Ah) × C-rating")
kv(35, "Full-throttle current (motors + electronics)", "=B19*B20+B32", "A")
kv(36, "Hover current per motor (approx)", "=B19*B24^1.5", "A", "current ≈ full × fraction^1.5")
kv(37, "Hover current total", "=B36*B20+B32", "A")
kv(38, "Hover power", "=B33*B37", "W")
ws["B38"].number_format = "0.0"
kv(39, "Estimated hover flight time", "=B29*B31/(B37*1000)*60", "min")
ws["B39"].number_format = "0.0"
kv(40, "Hover efficiency", f"={AUW}/B38", "g/W")
ws["B40"].number_format = "0.0"

# ---- 4) VERDICT ----
section(42, "4)  VERDICT / CHECKS")
verdicts = [
    ("Thrust-to-weight",
     '=IF(B22>=2,"PASS — controllable ("&TEXT(B22,"0.0")&":1)",'
     '"LOW ("&TEXT(B22,"0.0")&":1) — add thrust or cut weight")'),
    ("Battery can supply full throttle",
     '=IF(B34>=B35,"PASS — "&TEXT(B34-B35,"0.0")&" A margin",'
     '"FAIL — pack short by "&TEXT(B35-B34,"0.0")&" A; use higher C or bigger pack")'),
    ("Hover throttle headroom",
     '=IF(B24<=0.6,"PASS — "&TEXT(B24*100,"0")&"% hover throttle, good control margin",'
     '"HIGH — "&TEXT(B24*100,"0")&"% hover, little control margin")'),
    ("Estimated endurance", '=TEXT(B39,"0.0")&" min hover (less in wind / aggressive flight)"'),
]
for i, (lab, f) in enumerate(verdicts):
    r = 43 + i
    cell(f"A{r}", lab, font=BOLD, fill=VERD, bd=True)
    ws.merge_cells(f"B{r}:D{r}")
    cell(f"B{r}", f, fill=VERD, bd=True, align="left")

ws.freeze_panes = "A3"
wb.save("swarm-drone-calcs.xlsx")
print("swarm-drone-calcs.xlsx written")
